import openpyxl


class Racer:
    def __init__(self, comp_no, name, category, time):
        self.millisec = None
        self.comp_no = comp_no
        self.name = name
        self.category = category
        self.time = time

    def to_milli(self):
        minutes, sec, msec = map(int, self.time.split('.'))
        self.millisec = (minutes * 60 * 1000) + (sec * 1000) + msec
        return self.millisec

    def show_time(self):
        milli = self.millisec
        minutes = milli // (60 * 1000)
        milli %= (60 * 1000)
        seconds = milli // 1000
        milli %= 1000

        return f"{minutes}m {seconds}s {milli}ms"


def read_excel():
    wb = openpyxl.load_workbook('race_data.xlsx')
    ws = wb.active

    no_rows = ws.max_row - 1
    no_cols = ws.max_column

    # print(str(no_rows) + " " + str(ws.max_column))

    # for j in range(2, no_rows + 1):
    #     values = [ws.cell(row=j, column=i).value for i in range(1, no_cols + 1)]
    #     print(values)
    my_list = list()
    for cell in ws.iter_rows(min_row=2, max_row=no_rows + 1, min_col=1, max_col=no_cols):
        my_list.append(cell)

    return my_list


def get_racer_data():
    full_list = read_excel()
    racer_list = list()
    category_only = list()
    for comp_no, name, category, time in full_list:
        category_only.append(category.value.lower())
        racer_list.append(Racer(comp_no.value, name.value, category.value.lower(), time.value))

    # for racer in racer_list:
    #     print(racer.name, racer.to_milli())


    # print("\n\n")
    racer_sort_list = sorted(racer_list, key=lambda racer: racer.to_milli())
    # print(set(category_only))
    return racer_sort_list,set(category_only)

def data_by_category():
    racers, category = get_racer_data()

    category_dict = {}
    for racer in racers:

        racer_details = (racer.name, racer.show_time())
        if racer.category in category_dict:
            category_dict[racer.category].append(racer_details)
        else:
            category_dict[racer.category] = [racer_details]

    return category_dict

# get_category = data_by_category()
#
# print(get_category['open'])
